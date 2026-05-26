#!/usr/bin/env python3
"""Full 4-agent SEO audit pipeline test — tracks context/token statistics."""
import asyncio
import json
import sys
import time
import os
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, "/Users/maytinh/SEO Audit")

# ── Stats ─────────────────────────────────────────────────────────────────────
stats = {
    "start_time": time.time(),
    "pages_crawled": 0,
    "tools_called": 0,
    "tool_calls": [],
    "total_response_bytes": 0,
    "errors": [],
}

def log_tool(name, duration_ms, response_size_bytes, success=True):
    stats["tools_called"] += 1
    stats["total_response_bytes"] += response_size_bytes
    stats["tool_calls"].append({
        "tool": name,
        "ms": round(duration_ms),
        "kb": round(response_size_bytes / 1024, 1),
        "ok": success,
    })

def fmt(label, value): print(f"  {label:<38} {value}")

# ── Imports ───────────────────────────────────────────────────────────────────
from mcp_server.tools.crawler import crawl_page, check_url_batch
from mcp_server.tools.technical_checks import check_robots, check_sitemap
from mcp_server.checklist.technical import TECHNICAL_CHECKLIST, ChecklistItem
from mcp_server.checklist.ui import UI_CHECKLIST, UIChecklistItem
from mcp_server.analyzer.scorer import (
    ChecklistResult, CategoryScore, AuditScore,
    score_category, score_audit, get_top_issues, _compute_grade,
)

DOMAIN = "seongon.com"
BASE_URL = f"https://{DOMAIN}"

# ── Screaming Frog file (optional) ───────────────────────────────────────────
# Cách 1: Set đường dẫn trực tiếp ở đây
SF_FILE: str | None = None
# Cách 2: Truyền qua command line: python run_full_audit.py /path/to/sf_export.csv
if len(sys.argv) > 1:
    _sf_arg = Path(sys.argv[1])
    if _sf_arg.exists():
        SF_FILE = str(_sf_arg)
    else:
        print(f"  ⚠️  SF file không tồn tại: {sys.argv[1]}")

print(f"\n{'═'*62}")
print(f"  SEO AUDIT FULL PIPELINE — {DOMAIN}")
if SF_FILE:
    print(f"  📂 Screaming Frog: {Path(SF_FILE).name}")
print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(f"{'═'*62}")

# ════════════════════════════════════════════════════════════════════════════
# AGENT 1 — Config
# ════════════════════════════════════════════════════════════════════════════
audit_config = {
    "domain": DOMAIN,
    "brand_info": "Seongon - dịch vụ SEO, marketing online, agency SEO Việt Nam",
    "audit_purpose": "Audit toàn diện",
    "language": "Tiếng Việt",
    "pagespeed_api_key": None,
    "priority_groups": "Tất cả",
    "include_recommendations": True,
    "data_sources": {},
}

# ── Parse Screaming Frog data ─────────────────────────────────────────────────
sf_data: dict = {}
if SF_FILE:
    from mcp_server.tools.file_parsers import parse_screaming_frog as _parse_sf
    print(f"\n  📂 Đọc Screaming Frog: {SF_FILE}")
    sf_data = _parse_sf(SF_FILE)
    if "error" in sf_data:
        print(f"  ⚠️  SF parse error: {sf_data['error']}")
        sf_data = {}
    else:
        audit_config["data_sources"]["screaming_frog"] = SF_FILE
        n_sf = sf_data.get("total_urls", 0)
        n_err = len(sf_data.get("error_urls", []))
        n_dup = len(sf_data.get("duplicate_titles", []))
        n_noix = len(sf_data.get("non_indexable", []))
        _alt_raw = sf_data.get("images_missing_alt")
        n_noalt = len(_alt_raw) if _alt_raw is not None else None
        n_nometa = len(sf_data.get("missing_meta_descriptions", []))
        alt_str = f"{n_noalt} ảnh thiếu alt" if n_noalt is not None else "alt N/A (cần SF Images export)"
        print(f"  ✅ SF loaded: {n_sf:,} URLs | {n_err} lỗi 4xx/5xx | "
              f"{n_dup} nhóm title trùng | {n_noix} non-indexable | "
              f"{alt_str} | {n_nometa} thiếu meta desc")

print(f"\n✅ AGENT 1 — Config OK{' (có Screaming Frog)' if sf_data else ''}")

# ════════════════════════════════════════════════════════════════════════════
# AGENT 2 — Thu Thập Dữ Liệu
# ════════════════════════════════════════════════════════════════════════════
print(f"\n{'─'*62}\n🔍 AGENT 2 — Thu Thập Dữ Liệu\n{'─'*62}")

crawl_results = {}
robots_data = {}
sitemap_data = {}
batch_results_data = []

async def run_agent2():
    global crawl_results, robots_data, sitemap_data, batch_results_data

    # ── 2a: robots + sitemap parallel ────────────────────────────────────────
    print("\n[2a] robots.txt + sitemap.xml song song...")
    t0 = time.time()
    robots_data, sitemap_data = await asyncio.gather(
        check_robots(DOMAIN), check_sitemap(DOMAIN)
    )
    elapsed = (time.time() - t0) * 1000
    log_tool("seo_check_robots", elapsed / 2, len(json.dumps(robots_data)))
    log_tool("seo_check_sitemap", elapsed / 2, len(json.dumps(sitemap_data)))
    print(f"    robots: exists={robots_data.get('robots_exists')}, "
          f"sitemap_declared={bool(robots_data.get('sitemap_in_robots', []))}, "
          f"llms_txt={robots_data.get('has_llms_txt')}")
    print(f"    sitemap: exists={sitemap_data.get('sitemap_exists')}, "
          f"urls={sitemap_data.get('url_count', 0)}, "
          f"is_index={sitemap_data.get('is_index')}")

    # sitemap urls is list[dict] with "loc" key
    raw_sitemap = sitemap_data.get("urls", [])[:60]
    sitemap_urls = [u["loc"] if isinstance(u, dict) else u for u in raw_sitemap]

    # ── 2b: Crawl homepage ────────────────────────────────────────────────────
    print("\n[2b] Crawl 15 trang đại diện...")
    t0 = time.time()
    try:
        hp = await crawl_page(BASE_URL, page_type="homepage")
        elapsed_ms = (time.time() - t0) * 1000
        log_tool("crawl[homepage]", elapsed_ms, len(json.dumps(hp, ensure_ascii=False)))
        crawl_results["homepage"] = hp
        stats["pages_crawled"] += 1
        print(f"    ✅ [1/15] Homepage     — {hp.get('status_code')} — "
              f"title={hp.get('title_length')}ch — words={hp.get('word_count')}")
    except Exception as e:
        stats["errors"].append(f"homepage: {e}")
        print(f"    ❌ [1/15] Homepage     — {e}")
        hp = {}

    # Build candidate URL pool
    internal_hrefs = [lk["href"] for lk in hp.get("internal_links", [])
                      if lk.get("href", "").startswith("http")]
    candidates = list(dict.fromkeys(sitemap_urls + internal_hrefs))

    def pick_url(keywords, used):
        for u in candidates:
            ul = u.lower()
            if any(kw in ul for kw in keywords) and u not in used:
                return u
        for u in candidates:
            if u not in used:
                return u
        return None

    used = {BASE_URL}
    targets = [
        ("product_cat_large",  ["dich-vu", "service", "goi", "package", "seo-"],      "Danh mục DV lớn"),
        ("product_cat_small",  ["seo-website", "seo-local", "content", "audit"],      "Danh mục DV nhỏ"),
        ("product1",           ["/dich-vu/", "/service/", "goi-seo"],                 "Dịch vụ 1"),
        ("product2",           ["/dich-vu/", "/seo-"],                                "Dịch vụ 2"),
        ("blog_cat",           ["blog", "tin-tuc", "kien-thuc", "news"],              "Danh mục blog"),
        ("blog_cat2",          ["blog/", "kien-thuc/", "seo-co-ban"],                 "Danh mục blog 2"),
        ("article1",           ["/seo-", "/kien-thuc/", "/blog/"],                    "Bài viết 1"),
        ("article2",           ["/seo-", "/blog/", "/tin-tuc/"],                      "Bài viết 2"),
        ("about",              ["about", "gioi-thieu", "ve-chung-toi", "agency"],     "Về chúng tôi"),
        ("contact",            ["contact", "lien-he", "lien-lac"],                    "Liên hệ"),
        ("faq",                ["faq", "hoi-dap", "cau-hoi"],                         "FAQ"),
        ("static_page",        ["privacy", "terms", "chinh-sach", "dieu-khoan"],      "Trang tĩnh"),
        ("not_found",          None,                                                   "Trang 404"),
        ("sitemap_sample",     None,                                                   "URL sitemap"),
    ]

    for idx, (page_type, keywords, label) in enumerate(targets, 2):
        if page_type == "not_found":
            url = f"{BASE_URL}/trang-khong-ton-tai-xyz-404-test"
        elif page_type == "sitemap_sample":
            url = next((u for u in sitemap_urls[4:12] if u not in used), None)
            if not url:
                print(f"    ⏭️  [{idx}/15] {label} — không còn URL")
                continue
        else:
            url = pick_url(keywords, used) if keywords else None
            if not url:
                print(f"    ⏭️  [{idx}/15] {label} — không tìm được URL")
                continue

        used.add(url)
        t0 = time.time()
        try:
            result = await crawl_page(url, page_type=page_type)
            elapsed_ms = (time.time() - t0) * 1000
            log_tool(f"crawl[{page_type}]", elapsed_ms, len(json.dumps(result, ensure_ascii=False)))
            crawl_results[page_type] = result
            stats["pages_crawled"] += 1
            print(f"    ✅ [{idx}/15] {label:<18} — {result.get('status_code')} — "
                  f"title={result.get('title_length')}ch — {url[:55]}")
        except Exception as e:
            stats["errors"].append(f"{page_type}: {e}")
            print(f"    ❌ [{idx}/15] {label:<18} — {e}")

    # ── 2c: Batch URL check ───────────────────────────────────────────────────
    print("\n[2c] Batch URL check (20 URLs từ sitemap)...")
    batch_urls = sitemap_urls[:20]
    if batch_urls:
        t0 = time.time()
        try:
            batch_results_data = await check_url_batch(batch_urls)
            elapsed_ms = (time.time() - t0) * 1000
            log_tool("seo_check_url_batch", elapsed_ms, len(json.dumps(batch_results_data)))
            ok = sum(1 for r in batch_results_data if r.get("status_code") == 200)
            not_found = sum(1 for r in batch_results_data if r.get("status_code") == 404)
            redirects = sum(1 for r in batch_results_data if r.get("status_code") in [301, 302])
            print(f"    {len(batch_results_data)} URLs: ✅{ok} OK, ❌{not_found} 404, ↪{redirects} redirects")
        except Exception as e:
            stats["errors"].append(f"url_batch: {e}")
            print(f"    ❌ Batch lỗi: {e}")

asyncio.run(run_agent2())

hp = crawl_results.get("homepage", {})

print(f"\n  📊 Agent 2 tóm tắt: {stats['pages_crawled']} pages, "
      f"{stats['tools_called']} tools, {stats['total_response_bytes']/1024:.0f} KB")

# ════════════════════════════════════════════════════════════════════════════
# AGENT 3 — Phân Tích Checklist
# ════════════════════════════════════════════════════════════════════════════
print(f"\n{'─'*62}\n📋 AGENT 3 — Phân Tích Checklist\n{'─'*62}")
print(f"\n  Loaded: {len(TECHNICAL_CHECKLIST)} technical + {len(UI_CHECKLIST)} UI "
      f"= {len(TECHNICAL_CHECKLIST)+len(UI_CHECKLIST)} items")

def get_page(*page_types):
    for pt in page_types:
        if pt in crawl_results:
            return crawl_results[pt]
    return {}

def all_schema_types():
    types = []
    for data in crawl_results.values():
        types.extend(data.get("structured_data_types", []))
    return types

def check_dup_titles():
    titles = [d.get("title", "") for d in crawl_results.values() if d.get("title")]
    seen, dupes = set(), []
    for t in titles:
        if t in seen: dupes.append(t)
        seen.add(t)
    return dupes

def make_result(item, status, evidence, recommendation=None):
    return ChecklistResult(
        id=item.id, name=item.name, category=item.category,
        priority=item.priority, status=status, evidence=evidence,
        recommendation=recommendation or "",
    )

# ── Technical ─────────────────────────────────────────────────────────────────
tech_results = []

def map_tech(item):
    iid = item.id
    hp = crawl_results.get("homepage", {})

    # GROUP I — Domain
    if iid == "domain_01":
        rc = hp.get("redirect_chain", [])
        if rc:
            return "passed", f"Redirect chain: {' → '.join(rc[:3])}", None
        return "warning", "Không phát hiện redirect chain từ homepage", "Kiểm tra thủ công: curl -I http://www.tracdiahoangphat.com"

    if iid == "domain_02":
        url = hp.get("url", BASE_URL)
        if url.startswith("https://"):
            return "passed", f"HTTPS OK: {url}", None
        return "failed", "Website không dùng HTTPS", "Cài SSL và redirect HTTP → HTTPS toàn bộ"

    if iid == "domain_03":
        rt = hp.get("response_time_ms", 9999)
        if rt < 500: return "passed", f"Response time: {rt}ms (tốt)", None
        if rt < 1500: return "warning", f"Response time: {rt}ms (trung bình)", "Tối ưu server, CDN"
        return "failed", f"Response time: {rt}ms (chậm)", "Nâng cấp hosting, implement cache"

    # GROUP II — Indexability
    # index_01: robots.txt tồn tại và hợp lệ
    if iid == "index_01":
        if robots_data.get("robots_exists"):
            allowed = robots_data.get("googlebot_allowed", True)
            disallowed = robots_data.get("googlebot_disallowed_paths", [])
            status = "allow" if allowed else f"disallow: {disallowed[:2]}"
            return "passed", f"robots.txt tồn tại. Googlebot: {status}", None
        return "failed", "robots.txt không tồn tại", "Tạo robots.txt với User-agent: * và Sitemap:"

    # index_02: sitemap.xml tồn tại, hợp lệ, và khai báo trong robots.txt
    if iid == "index_02":
        sm_exists = sitemap_data.get("sitemap_exists")
        sitemap_in_robots = robots_data.get("sitemap_in_robots", [])
        sm_declared = sitemap_in_robots[0] if sitemap_in_robots else None
        n = sitemap_data.get("url_count", 0)
        if sm_exists and sm_declared:
            return "passed", f"sitemap.xml: {n} URLs, khai báo trong robots.txt: {sm_declared}", None
        if sm_exists:
            return "warning", f"sitemap.xml tồn tại ({n} URLs) nhưng chưa khai báo trong robots.txt", "Thêm 'Sitemap: https://domain.com/sitemap.xml' vào robots.txt"
        return "failed", "Không tìm thấy sitemap.xml", "Tạo XML sitemap, submit lên GSC, khai báo trong robots.txt"

    # index_03: Trang quan trọng được index (không bị noindex)
    if iid == "index_03":
        if sf_data and sf_data.get("non_indexable"):
            non_ix = sf_data["non_indexable"]
            total_sf = sf_data.get("total_urls", 0)
            reasons: dict[str, int] = {}
            for entry in non_ix:
                r = entry.get("reason", "unknown") or "unknown"
                reasons[r] = reasons.get(r, 0) + 1
            reason_str = ", ".join(f"{r}: {c}" for r, c in sorted(reasons.items(), key=lambda x: -x[1])[:4])
            sev = "failed" if len(non_ix) > 20 else "warning"
            return sev, f"[SF {total_sf:,} URLs] {len(non_ix)} trang non-indexable ({reason_str})", "Review danh sách trang noindex — xem xét loại bỏ noindex nếu muốn index"
        # fallback: crawl-based check
        important_pages = {k: v for k, v in crawl_results.items() if k != "not_found"}
        noindex_pages = []
        for page_key, data in important_pages.items():
            mr = (data.get("meta_robots") or "").lower()
            if "noindex" in mr:
                noindex_pages.append(f"{page_key}: {data.get('url','')[:50]}")
        if noindex_pages:
            return "failed", f"{len(noindex_pages)} trang quan trọng bị noindex: {noindex_pages[0]}", "Xóa noindex khỏi trang quan trọng"
        return "passed", f"Không có trang quan trọng nào bị noindex ({len(important_pages)} trang kiểm tra)", None

    # index_05: Không có trang 404 giả (soft 404) + broken links từ SF
    if iid == "index_05":
        if sf_data and sf_data.get("error_urls"):
            error_urls_sf = sf_data["error_urls"]
            total_sf = sf_data.get("total_urls", 0)
            codes: dict[str, int] = {}
            for u in error_urls_sf:
                c = str(u.get("status_code", "?"))
                codes[c] = codes.get(c, 0) + 1
            code_str = ", ".join(f"{c}: {n}" for c, n in sorted(codes.items()))
            sev = "failed" if len(error_urls_sf) > 5 else "warning"
            first_err = error_urls_sf[0].get("url", "")[:60] if error_urls_sf else ""
            return sev, f"[SF {total_sf:,} URLs] {len(error_urls_sf)} URLs lỗi ({code_str}). VD: {first_err}", "Sửa hoặc 301 redirect các URL trả về 4xx/5xx"
        # fallback: check crawled 404 page
        not_found = crawl_results.get("not_found", {})
        sc = not_found.get("status_code", 0)
        if sc == 404 or sc == 410:
            return "passed", f"Trang không tồn tại trả đúng {sc}", None
        if sc == 200:
            return "failed", f"Soft 404: URL không tồn tại trả về status {sc} (nội dung lỗi)", "Cấu hình server trả đúng 404/410 cho URL không tồn tại"
        return "manual", f"Không xác định được status code trang 404 (got {sc})", None

    # index_06: Canonical tag đúng và nhất quán
    if iid == "index_06":
        canonical = hp.get("canonical", "")
        ok = hp.get("canonical_matches_url", False)
        if canonical:
            st = "passed" if ok else "warning"
            rec = None if ok else "Đảm bảo canonical khớp URL thực tế"
            return st, f"Canonical: {canonical[:60]} ({'khớp' if ok else 'KHÔNG khớp'})", rec
        return "warning", "Không có canonical trên trang chủ", "Thêm <link rel='canonical'> cho tất cả trang"

    # GROUP III — JavaScript SEO
    if iid == "js_01":
        wc = hp.get("word_count", 0)
        if wc > 100: return "passed", f"Crawl không JS: {wc} từ — content OK", None
        if wc > 20: return "warning", f"Crawl không JS: {wc} từ — có thể ẩn qua JS", "Kiểm tra SSR/pre-rendering"
        return "failed", f"Crawl không JS: {wc} từ — phụ thuộc JS", "Implement SSR hoặc static rendering"

    # GROUP IV — Website Structure
    if iid == "struct_01":
        return "passed", f"URL trang chủ: {BASE_URL}", None

    if iid == "struct_02":
        from urllib.parse import urlparse
        violations = []
        warnings_list = []
        product_pages = ["product1", "product2"]
        article_pages = ["article1", "article2"]
        for pk in product_pages:
            data = crawl_results.get(pk, {})
            url = data.get("url", "")
            if not url:
                continue
            path = urlparse(url).path.strip("/")
            segments = [s for s in path.split("/") if s]
            if len(segments) < 2:
                violations.append(f"Sản phẩm ở root: {url}")
            elif len(segments) > 4:
                warnings_list.append(f"Sản phẩm quá sâu ({len(segments)} cấp): {url}")
        for pk in article_pages:
            data = crawl_results.get(pk, {})
            url = data.get("url", "")
            if not url:
                continue
            path = urlparse(url).path.strip("/")
            # strip .html extension before counting segments
            path = path.replace(".html", "").replace(".htm", "")
            segments = [s for s in path.split("/") if s]
            if len(segments) < 2:
                violations.append(f"Bài viết ở root: {url}")
            elif len(segments) > 4:
                warnings_list.append(f"Bài viết quá sâu ({len(segments)} cấp): {url}")
        if violations:
            return "failed", f"URL không đúng cấu trúc: {violations[0]}", "Di chuyển sản phẩm/bài viết vào danh mục: domain.com/danh-muc/san-pham"
        if warnings_list:
            return "warning", f"Cấu trúc URL hơi sâu: {warnings_list[0]}", "Cân nhắc giảm độ sâu URL xuống ≤3 cấp"
        checked = [crawl_results.get(p, {}).get("url", "") for p in product_pages + article_pages if crawl_results.get(p, {}).get("url")]
        if checked:
            return "passed", f"Cấu trúc URL hợp lý: {', '.join(checked[:2])}", None
        return "manual", "Không có đủ dữ liệu crawl product/article để kiểm tra", None

    if iid == "struct_03":
        # Breadcrumb không cần trên trang chủ — kiểm tra trên inner pages
        inner = ["product_cat_large", "product1", "article1", "blog_cat"]
        has_bc_inner = any(crawl_results.get(p, {}).get("has_breadcrumb_html", False) for p in inner)
        has_schema = "BreadcrumbList" in all_schema_types()
        if has_bc_inner or has_schema:
            bc_src = "inner pages" if has_bc_inner else "không có HTML"
            sc_src = "tìm thấy" if has_schema else "không có"
            return "passed", f"Breadcrumb: HTML={bc_src}, BreadcrumbList schema={sc_src}", None
        return "failed", "Không có breadcrumb HTML trên danh mục/sản phẩm hoặc BreadcrumbList schema", "Thêm breadcrumb nav + BreadcrumbList JSON-LD trên trang danh mục, sản phẩm"

    # struct_04: BreadcrumbList — check across all pages
    if iid == "struct_04":
        found = "BreadcrumbList" in all_schema_types()
        if found: return "passed", "BreadcrumbList schema: tìm thấy", None
        return "failed", "BreadcrumbList schema: KHÔNG tìm thấy", "Thêm JSON-LD BreadcrumbList trên trang danh mục/sản phẩm"

    # struct_05: Organization hoặc LocalBusiness
    if iid == "struct_05":
        types = all_schema_types()
        if "Organization" in types: return "passed", "Organization schema: tìm thấy", None
        if "LocalBusiness" in types: return "passed", "LocalBusiness schema: tìm thấy (thay thế Organization)", None
        return "warning", "Organization/LocalBusiness schema: KHÔNG tìm thấy", "Thêm JSON-LD Organization với name, url, logo, contactPoint"

    # struct_06: WebSite Schema
    if iid == "struct_06":
        found = "WebSite" in all_schema_types()
        if found: return "passed", "WebSite schema: tìm thấy", None
        return "warning", "WebSite schema: KHÔNG tìm thấy", "Thêm JSON-LD WebSite với potentialAction SearchAction"

    # struct_07: Product Schema — check on product pages
    if iid == "struct_07":
        product_types = []
        for p in ["product1", "product2"]:
            product_types.extend(crawl_results.get(p, {}).get("structured_data_types", []))
        if "Product" in product_types: return "passed", "Product schema: tìm thấy trên trang sản phẩm", None
        return "warning", "Product schema: KHÔNG tìm thấy trên trang sản phẩm", "Thêm JSON-LD Product với name, image, offers, price"

    # struct_08: Article hoặc BlogPosting — check on blog pages
    if iid == "struct_08":
        article_types = []
        for p in ["article1", "article2"]:
            article_types.extend(crawl_results.get(p, {}).get("structured_data_types", []))
        if "Article" in article_types or "BlogPosting" in article_types:
            found_type = "Article" if "Article" in article_types else "BlogPosting"
            return "passed", f"{found_type} schema: tìm thấy trên bài viết", None
        return "warning", "Article/BlogPosting schema: KHÔNG tìm thấy trên bài viết", "Thêm JSON-LD Article với headline, author, datePublished"

    # struct_09: FAQPage
    if iid == "struct_09":
        found = "FAQPage" in all_schema_types()
        if found: return "passed", "FAQPage schema: tìm thấy", None
        return "warning", "FAQPage schema: KHÔNG tìm thấy", "Thêm JSON-LD FAQPage nếu có trang FAQ"

    # struct_10: HowTo
    if iid == "struct_10":
        found = "HowTo" in all_schema_types()
        if found: return "passed", "HowTo schema: tìm thấy", None
        return "warning", "HowTo schema: KHÔNG tìm thấy", "Thêm JSON-LD HowTo cho trang hướng dẫn"

    # struct_11: Schema syntax check — validate bằng cách xem có types nào được extract không
    if iid == "struct_11":
        found_types = list(set(all_schema_types()))
        if found_types:
            return "passed", f"Phát hiện {len(found_types)} schema types: {', '.join(found_types[:6])}", None
        return "warning", "Không phát hiện JSON-LD schema nào — kiểm tra bằng schema.org/validator", "Thêm JSON-LD schemas và validate tại validator.schema.org"

    # struct_13: Pagination
    if iid == "struct_13":
        has_pag = any(crawl_results.get(p, {}).get("has_pagination", False) for p in ["product_cat_large", "blog_cat"])
        if has_pag: return "passed", "Phát hiện pagination trên trang danh mục/blog", None
        return "warning", "Không phát hiện pagination — kiểm tra thủ công nếu danh mục có nhiều trang", "Dùng số trang rõ ràng, tránh infinite scroll không crawlable"

    # GROUP V — Links
    if iid == "links_01":
        n = len(hp.get("internal_links", []))
        if n > 3: return "passed", f"Trang chủ: {n} internal links", None
        return "failed", f"Quá ít internal links ({n})", "Thêm internal links tăng crawlability"

    if iid == "links_02":
        no_nofollow = hp.get("external_links_without_nofollow", [])
        total_ext = len(hp.get("external_links", []))
        if no_nofollow:
            return "warning", f"{len(no_nofollow)}/{total_ext} external links thiếu nofollow", "Thêm rel='nofollow' cho external links"
        return "passed", f"Tất cả {total_ext} external links có nofollow hoặc không có", None

    # GROUP VI — Metadata
    if iid == "meta_01":
        title = hp.get("title", "")
        tl = hp.get("title_length", 0)
        if not title: return "failed", "Không có title", "Thêm <title> 50-65 ký tự với từ khóa chính"
        if tl < 30: return "failed", f"Title quá ngắn: '{title}' ({tl}ch)", "Viết title 50-65 ký tự"
        if tl > 65: return "warning", f"Title quá dài: '{title[:50]}...' ({tl}ch)", "Rút ngắn title ≤65 ký tự"
        return "passed", f"Title OK: '{title[:60]}' ({tl}ch)", None

    # meta_02: Meta description length (NOT duplicate check)
    if iid == "meta_02":
        md = hp.get("meta_description", "")
        ml = hp.get("meta_description_length", 0)
        if not md: return "failed", "Không có meta description", "Thêm meta description 120-160 ký tự"
        if ml < 80: return "warning", f"Meta desc quá ngắn ({ml}ch)", "Mở rộng lên 120-160 ký tự"
        if ml > 160: return "warning", f"Meta desc quá dài ({ml}ch), bị cắt SERP", "Rút ngắn ≤160 ký tự"
        return "passed", f"Meta description OK: {ml}ch", None

    # meta_03: Duplicate title/meta description check (cross-page)
    if iid == "meta_03":
        if sf_data:
            dup_titles_sf = sf_data.get("duplicate_titles", [])
            missing_titles_sf = sf_data.get("missing_titles", [])
            missing_meta_sf = sf_data.get("missing_meta_descriptions", [])
            missing_h1_sf = sf_data.get("missing_h1", [])
            total_sf = sf_data.get("total_urls", 0)
            issues = []
            if dup_titles_sf:
                first = dup_titles_sf[0]
                issues.append(f"{len(dup_titles_sf)} nhóm title trùng (VD: '{first['title'][:40]}' — {len(first['urls'])} trang)")
            if missing_titles_sf:
                issues.append(f"{len(missing_titles_sf)}/{total_sf} trang thiếu title")
            if missing_meta_sf:
                issues.append(f"{len(missing_meta_sf)}/{total_sf} trang thiếu meta description")
            if missing_h1_sf:
                issues.append(f"{len(missing_h1_sf)}/{total_sf} trang thiếu H1")
            if issues:
                sev = "failed" if (len(dup_titles_sf) > 10 or len(missing_titles_sf) > 5) else "warning"
                return sev, f"[SF {total_sf:,} URLs] " + "; ".join(issues), "Đảm bảo mỗi trang có title unique 50-65ch, meta desc unique 120-160ch, 1 H1"
            return "passed", f"[SF {total_sf:,} URLs] Không phát hiện duplicate title hoặc thiếu meta desc/H1", None
        # fallback: 15-page crawl comparison
        title_map: dict[str, list[str]] = {}
        desc_map: dict[str, list[str]] = {}
        for page_key, data in crawl_results.items():
            t = (data.get("title") or "").strip()
            d = (data.get("meta_description") or "").strip()
            if t: title_map.setdefault(t, []).append(page_key)
            if d: desc_map.setdefault(d, []).append(page_key)
        dup_titles = {t: pages for t, pages in title_map.items() if len(pages) > 1}
        dup_descs = {d: pages for d, pages in desc_map.items() if len(pages) > 1}
        issues = []
        if dup_titles:
            first = next(iter(dup_titles.items()))
            issues.append(f"{len(dup_titles)} title trùng (VD: '{first[0][:40]}' trên {first[1]})")
        if dup_descs:
            first = next(iter(dup_descs.items()))
            issues.append(f"{len(dup_descs)} meta desc trùng (VD: '{first[0][:40]}...' trên {first[1]})")
        if issues:
            return "warning", "; ".join(issues), "Đảm bảo mỗi trang có title và meta description unique"
        return "passed", f"Không phát hiện title/meta desc trùng trong {stats['pages_crawled']} trang (chưa có SF — độ chính xác thấp)", None

    # meta_04: H1
    if iid == "meta_04":
        h1s = hp.get("headings", {}).get("h1", [])
        cnt = hp.get("h1_count", 0)
        if cnt == 1: return "passed", f"H1: '{h1s[0][:60]}'", None
        if cnt == 0: return "failed", "Không có H1 trên trang chủ", "Thêm 1 thẻ H1 chứa từ khóa chính"
        return "warning", f"{cnt} thẻ H1 (nên chỉ 1): {[h[:30] for h in h1s[:2]]}", "Dùng đúng 1 thẻ H1/trang"

    # meta_05: Cấu trúc heading
    if iid == "meta_05":
        headings = hp.get("headings", {})
        h2s = headings.get("h2", [])
        h3s = headings.get("h3", [])
        if h2s: return "passed", f"Heading: H1={hp.get('h1_count',0)}, H2={len(h2s)}, H3={len(h3s)}", None
        return "warning", "Không có H2 — cấu trúc heading chưa tốt", "Thêm H2 để phân chia nội dung"

    # meta_06: Alt text hình ảnh
    if iid == "meta_06":
        # SF Images tab export có alt_text + image_src → images_missing_alt là list
        # SF All Crawl Data export không có cột đó → images_missing_alt là None
        if sf_data and sf_data.get("images_missing_alt") is not None:
            no_alt_sf = sf_data["images_missing_alt"]
            total_sf = sf_data.get("total_urls", 0)
            if not no_alt_sf:
                return "passed", f"[SF Images] Không phát hiện ảnh thiếu alt text", None
            sev = "failed" if len(no_alt_sf) > 10 else "warning"
            sample = no_alt_sf[0][:60] if no_alt_sf else ""
            return sev, f"[SF Images] {len(no_alt_sf)} ảnh thiếu alt text. VD: {sample}", "Thêm alt text mô tả nội dung ảnh, chứa từ khóa liên quan cho tất cả ảnh"
        # fallback: dùng crawl data (15 trang, homepage + các trang crawled)
        no_alt = hp.get("images_without_alt", [])
        total_imgs = len(hp.get("images", []))
        note = " (chưa có SF Images export — chỉ kiểm tra trang chủ)" if sf_data else " (chưa có SF)"
        if not no_alt: return "passed", f"Tất cả {total_imgs} ảnh trang chủ có alt text{note}", None
        st = "warning" if len(no_alt) <= 3 else "failed"
        return st, f"{len(no_alt)}/{total_imgs} ảnh trang chủ thiếu alt text{note}", "Thêm alt text cho tất cả ảnh"

    # meta_07: Open Graph tags
    if iid == "meta_07":
        og = hp.get("og_tags", {})
        has_title = bool(og.get("og:title"))
        has_desc = bool(og.get("og:description"))
        has_img = bool(og.get("og:image"))
        if has_title and has_desc and has_img:
            return "passed", "OG: title✅ description✅ image✅", None
        missing = [k for k, v in [("og:title", has_title), ("og:description", has_desc), ("og:image", has_img)] if not v]
        if missing:
            return "failed" if not has_title else "warning", f"Thiếu OG: {', '.join(missing)}", f"Thêm {', '.join(missing)}"
        return "failed", "Không có Open Graph tags", "Thêm og:title, og:description, og:image, og:url"

    # meta_08: Viewport meta tag
    if iid == "meta_08":
        vp = hp.get("viewport_meta", "")
        if vp and "width=device-width" in vp: return "passed", f"Viewport: {vp}", None
        if vp: return "warning", f"Viewport bất thường: {vp}", "Dùng: width=device-width, initial-scale=1"
        return "failed", "Không có viewport meta", "Thêm <meta name='viewport' content='width=device-width, initial-scale=1'>"

    # meta_09: Charset
    if iid == "meta_09":
        cs = hp.get("charset", "")
        if cs and "utf" in cs.lower(): return "passed", f"Charset: {cs}", None
        if cs: return "warning", f"Charset: {cs} (khuyến nghị UTF-8)", None
        return "warning", "Không phát hiện charset", "Thêm <meta charset='UTF-8'>"

    # meta_10: Hreflang
    if iid == "meta_10":
        hl = hp.get("hreflang", [])
        if hl: return "passed", f"Hreflang: {hl}", None
        return "manual", "Không có hreflang — kiểm tra nếu website đa ngôn ngữ", None

    # meta_11: Favicon
    if iid == "meta_11":
        fav = hp.get("favicon", "")
        if fav: return "passed", f"Favicon: {fav}", None
        return "failed", "Không tìm thấy favicon", "Thêm favicon.ico và <link rel='icon'>"

    if iid == "meta_12":
        # Rich snippet (Review/Rating) chỉ check trên trang sản phẩm và bài viết, không cần trên homepage
        rich_product = get_page("product1", "product2").get("has_rich_snippet_markup", False)
        rich_article = get_page("article1", "article2").get("has_rich_snippet_markup", False)
        if rich_product:
            return "passed", "Trang sản phẩm có rich snippet markup (Review/Rating schema)", None
        if rich_article:
            return "warning", "Bài viết có rich snippet nhưng trang sản phẩm chưa có", "Thêm schema AggregateRating cho trang sản phẩm"
        return "warning", "Không phát hiện Review/Rating schema trên trang sản phẩm hoặc bài viết", "Thêm JSON-LD AggregateRating/Review cho trang sản phẩm"

    if iid == "meta_13":
        has_toc = get_page("article1", "article2").get("has_toc", False)
        if has_toc: return "passed", "Bài viết có Table of Contents", None
        return "warning", "Bài viết không có TOC", "Thêm TOC cho bài viết dài >1000 từ"

    # GROUP VII — UI & Responsive
    if iid == "ui_02":
        return "manual", "Kiểm tra thủ công: Google Mobile-Friendly Test hoặc Chrome DevTools", None

    # GROUP VIII — Features & UX
    if iid == "feat_01":
        hs = hp.get("has_search_form", False)
        if hs: return "passed", "Có form tìm kiếm nội bộ", None
        return "warning", "Không phát hiện internal search", "Thêm search form (đặc biệt với site thương mại)"

    if iid == "feat_02":
        hc = get_page("contact", "homepage").get("has_contact_form", False)
        if hc: return "passed", "Có form liên hệ", None
        return "warning", "Không phát hiện contact form", "Thêm form liên hệ để tăng conversion"

    if iid == "feat_03":
        hlc = hp.get("has_live_chat", False)
        plat = hp.get("live_chat_platform", "")
        if hlc: return "passed", f"Live chat: {plat or 'phát hiện được'}", None
        return "warning", "Không có live chat", "Cân nhắc Tawk.to (miễn phí)"

    if iid == "feat_04":
        hs = get_page("article1", "article2", "homepage").get("has_social_share_buttons", False)
        if hs: return "passed", "Có nút chia sẻ mạng xã hội", None
        return "warning", "Không phát hiện social share buttons", "Thêm share buttons cho bài viết và sản phẩm"

    if iid == "feat_05":
        hcm = get_page("article1", "article2").get("has_comment_section", False)
        if hcm: return "passed", "Có phần bình luận", None
        return "warning", "Không phát hiện comment section", "Cân nhắc thêm bình luận để tăng engagement"

    if iid == "feat_06":
        hn = hp.get("has_newsletter_form", False)
        if hn: return "passed", "Có newsletter form", None
        return "warning", "Không có newsletter form", "Thêm newsletter để build email list"

    if iid == "feat_07":
        hcart = get_page("product1", "product2", "homepage").get("has_cart_button", False)
        hprice = get_page("product1", "product2").get("has_price_display", False)
        if hcart and hprice: return "passed", "Có nút mua hàng và giá sản phẩm", None
        if hprice: return "warning", "Hiển thị giá nhưng thiếu nút mua hàng rõ", "Thêm CTA (Thêm vào giỏ / Mua ngay)"
        return "manual", "Không xác định được e-commerce features — kiểm tra thủ công", None

    if iid == "feat_08":
        hb2t = hp.get("has_back_to_top", False)
        if hb2t: return "passed", "Có Back to Top button", None
        return "warning", "Không có Back to Top", "Thêm Back to Top button cho UX trang dài"

    # GROUP IX — Page Speed
    if iid == "speed_01":
        cc = hp.get("cache_control", "")
        if cc and cc != "no-store":
            return "passed", f"Cache-Control: {cc}", None
        return "failed", f"Cache-Control: {cc or '(không có)'}", "Cài Cache-Control headers, max-age cho static assets"

    if iid == "speed_02":
        rt = hp.get("response_time_ms", 9999)
        if rt < 200: return "passed", f"Response time: {rt}ms (xuất sắc)", None
        if rt < 500: return "passed", f"Response time: {rt}ms (tốt)", None
        if rt < 1500: return "warning", f"Response time: {rt}ms (trung bình)", "Enable gzip, caching, CDN"
        return "failed", f"Response time: {rt}ms (chậm)", "Nâng cấp hosting, Redis cache, CDN"

    if iid in ("speed_03", "speed_04"):
        return "manual", "Cần PageSpeed Insights API key để đo Core Web Vitals", None

    # GROUP X — CMS & Technical
    # cms_14: llms.txt — auto-check via robots data
    if iid == "cms_14":
        if robots_data.get("has_llms_txt"):
            return "passed", f"llms.txt tồn tại: {robots_data.get('llms_txt_url', '')}", None
        return "warning", "llms.txt chưa có", "Tạo /llms.txt để AI crawlers (ChatGPT, Perplexity, Claude) hiểu website tốt hơn. Xem llmstxt.org"

    # GROUP XI — Measurement
    if iid == "measure_01":
        hgtm = hp.get("has_gtm", False)
        gtm_id = hp.get("gtm_id", "")
        if hgtm: return "passed", f"GTM: {gtm_id or 'phát hiện được'}", None
        return "warning", "Không phát hiện Google Tag Manager", "Cài GTM để quản lý tracking tập trung"

    if iid == "measure_02":
        hga4 = hp.get("has_ga4", False)
        ga4_id = hp.get("ga4_id", "")
        if hga4: return "passed", f"GA4: {ga4_id or 'phát hiện được'}", None
        return "failed", "Không phát hiện Google Analytics 4", "Cài đặt GA4 để theo dõi traffic"

    # GROUP XIV — Log File
    if iid.startswith("log_"):
        return "manual", "Cần file access log từ server", None

    return "manual", "Chưa có logic map cho tiêu chí này", None


print("\n  [3a] Mapping technical checklist...")
for item in TECHNICAL_CHECKLIST:
    status, evidence, rec = map_tech(item)
    tech_results.append(make_result(item, status, evidence, rec if audit_config["include_recommendations"] else None))

# ── UI Checklist ──────────────────────────────────────────────────────────────
print("  [3b] Mapping UI checklist...")
ui_results = []

page_type_map = {
    "header": "homepage",
    "footer": "homepage",
    "homepage": "homepage",
    "product_category": "product_cat_large",
    "product_detail": "product1",
    "blog_category": "blog_cat",
    "article": "article1",
    "about": "about",
    "contact": "contact",
    "404": "not_found",
}

for item in UI_CHECKLIST:
    ui_category = f"UI — {item.page_type.replace('_', ' ').title()}"
    iid = item.id

    if item.check_mode == "auto" and item.crawl_field:
        crawl_key = page_type_map.get(item.page_type, "homepage")
        page_data = crawl_results.get(crawl_key, hp)

        # ── Special-case items ──────────────────────────────────────────
        if iid == "home_10":  # popup: có = warning, không có = passed
            if page_data.get("has_popup"):
                status = "warning"
                ev = "Phát hiện popup/modal trong HTML — kiểm tra không che khuất nội dung chính"
                rec = "Đảm bảo popup có nút đóng rõ ràng, không che full-screen, không autoplay trên mobile"
            else:
                status, ev, rec = "passed", "Không phát hiện popup che khuất nội dung", None

        elif iid == "home_13":  # tốc độ = proxy qua server response time
            rt = page_data.get("response_time_ms", 0)
            if rt < 800:
                status, ev, rec = "passed", f"Server response time: {rt}ms", None
            elif rt < 2000:
                status, ev, rec = "warning", f"Server response time: {rt}ms — nên đo LCP bằng PageSpeed Insights", "Dùng https://pagespeed.web.dev để đo LCP, FCP trên mobile"
            else:
                status, ev, rec = "failed", f"Server response time: {rt}ms (chậm)", "Tối ưu TTFB: CDN, server-side cache, giảm blocking requests"

        elif iid == "about_02":  # câu chuyện thương hiệu — proxy word_count
            wc = page_data.get("word_count", 0)
            if wc >= 300:
                status, ev, rec = "passed", f"Trang About có {wc} từ — đủ nội dung câu chuyện thương hiệu", None
            else:
                status, ev, rec = "warning", f"Trang About chỉ có {wc} từ — có thể thiếu nội dung", "Thêm: lịch sử hình thành, founder story, hành trình phát triển (tối thiểu 300 từ)"

        elif iid == "notfound_02":  # thông báo lỗi thân thiện — proxy word_count
            wc = page_data.get("word_count", 0)
            if wc >= 50:
                status, ev, rec = "passed", f"Trang 404 có {wc} từ nội dung", None
            else:
                status, ev, rec = "failed", f"Trang 404 quá ít nội dung ({wc} từ)", "Thiết kế trang 404 thân thiện: thông báo rõ ràng, link trang chủ, ô tìm kiếm"

        elif iid == "notfound_03":  # link về trang chủ
            home_url = BASE_URL.rstrip("/")
            links = page_data.get("internal_links", [])
            has_home = any(
                l.get("href", "").rstrip("/") == home_url or l.get("href", "") == "/"
                for l in links
            )
            if has_home:
                status, ev, rec = "passed", "Trang 404 có link về trang chủ", None
            else:
                status, ev, rec = "failed", "Không tìm thấy link về trang chủ trên trang 404", "Thêm button 'Về trang chủ' nổi bật"

        elif iid == "notfound_05":  # gợi ý trang phổ biến
            n_links = len(page_data.get("internal_links", []))
            if n_links > 8:
                status, ev, rec = "passed", f"Trang 404 có {n_links} internal links — đủ điều hướng", None
            else:
                status, ev, rec = "warning", f"Trang 404 chỉ có {n_links} internal links", "Thêm section 'Gợi ý trang': trang chủ, danh mục chính, liên hệ"

        elif iid == "notfound_07":  # header + footer đầy đủ trên 404
            has_nav = page_data.get("header_has_nav", False)
            has_footer = page_data.get("footer_has_copyright", False)
            if has_nav and has_footer:
                status, ev, rec = "passed", "Trang 404 có header nav và footer đầy đủ", None
            elif has_nav:
                status, ev, rec = "warning", "Trang 404 có header nav nhưng thiếu footer", "Dùng chung template với trang thường để footer xuất hiện"
            else:
                status, ev, rec = "failed", "Trang 404 thiếu header nav và/hoặc footer", "Đảm bảo trang 404 dùng chung layout với các trang khác"

        elif iid == "blog_cat_02":  # card bài viết có ảnh + heading
            imgs = page_data.get("images", [])
            h23 = len(page_data.get("headings", {}).get("h2", [])) + len(page_data.get("headings", {}).get("h3", []))
            if len(imgs) >= 3 and h23 >= 3:
                status, ev, rec = "passed", f"Blog listing: {len(imgs)} ảnh, {h23} heading H2/H3 — card đủ thành phần", None
            elif len(imgs) >= 1:
                status, ev, rec = "warning", f"Blog listing: {len(imgs)} ảnh, {h23} heading — kiểm tra mỗi card đủ thumbnail+tiêu đề+excerpt", "Đảm bảo mỗi bài có featured image và excerpt"
            else:
                status, ev, rec = "failed", "Blog listing không có ảnh thumbnail", "Thêm featured image cho tất cả bài viết"

        else:
            # ── Default boolean / value logic ───────────────────────────
            val = page_data.get(item.crawl_field)
            if val is True:
                status, ev, rec = "passed", f"{item.crawl_field}: có", None
            elif val is False:
                status, ev, rec = "failed", f"{item.crawl_field}: không có", f"Thêm: {item.name}"
            elif isinstance(val, list):
                if val:
                    status, ev, rec = "passed", f"{item.crawl_field}: {len(val)} items", None
                else:
                    status, ev, rec = "failed", f"{item.crawl_field}: danh sách rỗng", f"Thêm: {item.name}"
            elif isinstance(val, dict):
                if val:
                    status, ev, rec = "passed", f"{item.crawl_field}: có dữ liệu", None
                else:
                    status, ev, rec = "failed", f"{item.crawl_field}: rỗng", f"Thêm: {item.name}"
            elif isinstance(val, int):
                if val > 0:
                    status, ev, rec = "passed", f"{item.crawl_field}={val}", None
                else:
                    status, ev, rec = "failed", f"{item.crawl_field}=0", f"Thêm: {item.name}"
            elif isinstance(val, str) and val:
                status, ev, rec = "passed", f"{item.crawl_field}: {val[:80]}", None
            elif val is None:
                status, ev, rec = "manual", f"Field '{item.crawl_field}' không có trong crawl data", None
            else:
                status, ev, rec = "failed", f"{item.crawl_field}: không có hoặc rỗng", f"Thêm: {item.name}"
    elif item.check_mode == "screenshot":
        status = "manual"
        ev = f"[Cần screenshot] {item.manual_guide or 'Chụp screenshot và paste vào chat để Claude phân tích'}"
        rec = None
    else:
        status = "manual"
        ev = item.manual_guide or "Cần kiểm tra thủ công"
        rec = None

    rec_val = rec if audit_config["include_recommendations"] else None
    ui_results.append(ChecklistResult(
        id=item.id, name=item.name, category=ui_category,
        priority=item.priority, status=status, evidence=ev,
        recommendation=rec_val or "",
    ))

# ── Score ─────────────────────────────────────────────────────────────────────
all_results = tech_results + ui_results

results_by_category: dict[str, list[ChecklistResult]] = {}
for r in all_results:
    results_by_category.setdefault(r.category, []).append(r)

audit_score = score_audit(results_by_category)
top_issues = get_top_issues(results_by_category, n=10)

print(f"  ✅ Mapping xong: {len(all_results)} tiêu chí đánh giá")

# ════════════════════════════════════════════════════════════════════════════
# AGENT 4 — Xuất Báo Cáo
# ════════════════════════════════════════════════════════════════════════════
print(f"\n{'─'*62}\n📄 AGENT 4 — Xuất Báo Cáo\n{'─'*62}")

# Build report dict compatible with server's _handle_save_report
report_dir = Path("/Users/maytinh/SEO Audit/reports")
report_dir.mkdir(parents=True, exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
filename = f"seo_report_{DOMAIN.replace('.', '_')}_{timestamp}.md"
filepath = report_dir / filename

lines = []
lines += [
    f"# Báo Cáo SEO Audit — {DOMAIN}",
    f"",
    f"**Ngày tạo:** {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    f"**Domain:** {BASE_URL}",
    f"**Thương hiệu:** {audit_config['brand_info']}",
    f"**Mục đích:** {audit_config['audit_purpose']}",
]
if sf_data:
    n_sf = sf_data.get("total_urls", 0)
    lines.append(f"**Screaming Frog:** {n_sf:,} URLs crawled | "
                 f"{len(sf_data.get('error_urls', []))} lỗi 4xx/5xx | "
                 f"{len(sf_data.get('duplicate_titles', []))} nhóm title trùng | "
                 f"{len(sf_data.get('missing_meta_descriptions', []))} thiếu meta desc | "
                 f"{len(sf_data.get('images_missing_alt') or []) or 'N/A (cần SF Images export)'} ảnh thiếu alt")
else:
    lines.append(f"**Screaming Frog:** Không có — phân tích dựa trên crawl 15 trang")
lines += [
    f"",
    f"---",
    f"",
    f"## Tổng Điểm: {audit_score.percentage:.1f}% — Grade **{audit_score.grade}**",
    f"",
    f"| Trạng thái | Số tiêu chí |",
    f"|------------|-------------|",
    f"| ✅ Đạt | {audit_score.passed} |",
    f"| ❌ Lỗi | {audit_score.failed} |",
    f"| ⚠️ Cần cải thiện | {audit_score.warning} |",
    f"| 🔍 Cần kiểm tra thủ công | {audit_score.manual} |",
    f"| **Tổng** | **{len(all_results)}** |",
    f"",
    f"---",
    f"",
    f"## Điểm Theo Nhóm",
    f"",
    f"| Nhóm | Điểm | Grade | ✅ | ❌ | ⚠️ | 🔍 |",
    f"|------|------|-------|---|---|---|---|",
]
for cat, results_list in sorted(results_by_category.items()):
    cs = score_category(results_list)
    lines.append(f"| {cat} | {cs.percentage:.0f}% | {cs.grade} | {cs.passed} | {cs.failed} | {cs.warning} | {cs.manual} |")

lines += [
    f"",
    f"---",
    f"",
    f"## Top 10 Vấn Đề Ưu Tiên",
    f"",
]
for i, r in enumerate(top_issues, 1):
    icon = "❌" if r.status == "failed" else "⚠️"
    lines.append(f"### {i}. {icon} {r.name}")
    lines.append(f"- **Nhóm:** {r.category}  |  **Độ ưu tiên:** {r.priority}")
    lines.append(f"- **Bằng chứng:** {r.evidence}")
    if r.recommendation:
        lines.append(f"- **Đề xuất:** {r.recommendation}")
    lines.append("")

lines += [f"---", f"", f"## Chi Tiết Toàn Bộ Tiêu Chí", f""]

current_cat = ""
for r in all_results:
    if r.category != current_cat:
        current_cat = r.category
        lines += [f"", f"### {current_cat}", f"", f"| Tiêu chí | Trạng thái | Bằng chứng | Đề xuất |", f"|----------|------------|------------|---------|"]
    icon = {"passed": "✅", "failed": "❌", "warning": "⚠️", "manual": "🔍"}.get(r.status, "?")
    ev = (r.evidence or "")[:80].replace("|", "\\|")
    rec = (r.recommendation or "")[:80].replace("|", "\\|")
    lines.append(f"| {r.name} | {icon} {r.status} | {ev} | {rec} |")

filepath.write_text("\n".join(lines), encoding="utf-8")
log_tool("seo_save_report", 50, len("\n".join(lines)))
print(f"\n  ✅ Báo cáo MD: {filepath}")

# ── Excel report ─────────────────────────────────────────────────────────────
def _col_width(ws, col, width): ws.column_dimensions[get_column_letter(col)].width = width

def _header_row(ws, headers, fill_hex="1F6FEB"):
    fill = PatternFill("solid", fgColor=fill_hex)
    bold = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold; cell.fill = fill; cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

STATUS_COLORS = {
    "passed":  "D6F5D6",  # xanh lá nhạt
    "failed":  "FADADD",  # đỏ nhạt
    "warning": "FFF3CD",  # vàng nhạt
    "manual":  "E8E8E8",  # xám nhạt
}
STATUS_LABELS = {"passed": "✅ Đạt", "failed": "❌ Lỗi", "warning": "⚠️ Cảnh báo", "manual": "🔍 Thủ công"}

wb = openpyxl.Workbook()

# ─── Sheet 1: Tóm tắt ────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Tóm tắt"
ws1.append(["SEO AUDIT REPORT"])
ws1["A1"].font = Font(bold=True, size=16)
ws1.merge_cells("A1:D1")
ws1.append([])
ws1.append(["Domain", BASE_URL])
ws1.append(["Ngày audit", datetime.now().strftime("%d/%m/%Y %H:%M")])
ws1.append(["Thương hiệu", audit_config["brand_info"]])
ws1.append(["Tổng điểm", f"{audit_score.percentage:.1f}%"])
ws1.append(["Grade", audit_score.grade])
ws1.append([])
ws1.append(["Trạng thái", "Số tiêu chí"])
for row in [("✅ Đạt", audit_score.passed), ("❌ Lỗi", audit_score.failed),
            ("⚠️ Cảnh báo", audit_score.warning), ("🔍 Thủ công", audit_score.manual),
            ("Tổng", len(all_results))]:
    ws1.append(list(row))
ws1.append([])
ws1.append(["Top 10 Vấn Đề Ưu Tiên"])
ws1.cell(ws1.max_row, 1).font = Font(bold=True, size=12)
ws1.append(["#", "Nhóm", "Tiêu chí", "Trạng thái", "Bằng chứng", "Đề xuất"])
for c in range(1, 7):
    ws1.cell(ws1.max_row, c).font = Font(bold=True)
for i, r in enumerate(top_issues, 1):
    ws1.append([i, r.category, r.name, STATUS_LABELS.get(r.status, r.status), r.evidence, r.recommendation])
    fill = PatternFill("solid", fgColor=STATUS_COLORS.get(r.status, "FFFFFF"))
    ws1.cell(ws1.max_row, 4).fill = fill
_col_width(ws1, 1, 5); _col_width(ws1, 2, 28); _col_width(ws1, 3, 45)
_col_width(ws1, 4, 16); _col_width(ws1, 5, 55); _col_width(ws1, 6, 50)

# ─── Sheet 2: Chi tiết ────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Chi tiết")
hdrs = ["ID", "Nhóm", "Tiêu chí", "Ưu tiên", "Trạng thái", "Bằng chứng", "Đề xuất"]
_header_row(ws2, hdrs)
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
for r in all_results:
    row = [r.id, r.category, r.name, r.priority,
           STATUS_LABELS.get(r.status, r.status), r.evidence, r.recommendation]
    ws2.append(row)
    fill = PatternFill("solid", fgColor=STATUS_COLORS.get(r.status, "FFFFFF"))
    for c in range(1, 8):
        cell = ws2.cell(ws2.max_row, c)
        cell.fill = fill; cell.border = thin
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:G{len(all_results)+1}"
_col_width(ws2, 1, 14); _col_width(ws2, 2, 30); _col_width(ws2, 3, 45)
_col_width(ws2, 4, 12); _col_width(ws2, 5, 14); _col_width(ws2, 6, 60); _col_width(ws2, 7, 55)

# ─── Sheet 3: Điểm theo nhóm ─────────────────────────────────────────────────
ws3 = wb.create_sheet("Điểm theo nhóm")
_header_row(ws3, ["Nhóm", "Điểm %", "Grade", "✅ Đạt", "❌ Lỗi", "⚠️ Cảnh báo", "🔍 Thủ công"])
for cat, rlist in sorted(results_by_category.items()):
    cs = score_category(rlist)
    ws3.append([cat, round(cs.percentage, 1), cs.grade, cs.passed, cs.failed, cs.warning, cs.manual])
    cell = ws3.cell(ws3.max_row, 3)
    grade_color = {"A": "D6F5D6", "B": "E8F5E9", "C": "FFF9C4", "D": "FFE0B2", "F": "FADADD"}.get(cs.grade, "FFFFFF")
    cell.fill = PatternFill("solid", fgColor=grade_color)
ws3.freeze_panes = "A2"
_col_width(ws3, 1, 35); _col_width(ws3, 2, 10); _col_width(ws3, 3, 8)
_col_width(ws3, 4, 10); _col_width(ws3, 5, 10); _col_width(ws3, 6, 14); _col_width(ws3, 7, 14)

# ─── Sheet 4: Screaming Frog (nếu có) ───────────────────────────────────────
if sf_data:
    ws_sf = wb.create_sheet("Screaming Frog")
    _header_row(ws_sf, ["Loại", "Số lượng", "Chi tiết (50 đầu)"], fill_hex="7B68EE")
    _col_width(ws_sf, 1, 30); _col_width(ws_sf, 2, 12); _col_width(ws_sf, 3, 80)

    def _sf_row(ws, label, count, detail=""):
        ws.append([label, count, detail])
        ws.cell(ws.max_row, 2).alignment = Alignment(horizontal="center")
        ws.cell(ws.max_row, 3).alignment = Alignment(wrap_text=True, vertical="top")

    _sf_row(ws_sf, "Tổng URLs crawled", sf_data.get("total_urls", 0))

    # Status code breakdown
    sc_bd = sf_data.get("status_code_breakdown", {})
    for code, cnt in sorted(sc_bd.items(), key=lambda x: str(x[0])):
        _sf_row(ws_sf, f"Status {code}", cnt)

    # Errors
    err_urls = sf_data.get("error_urls", [])
    if err_urls:
        _sf_row(ws_sf, "URLs lỗi 4xx/5xx", len(err_urls),
                " | ".join(f"{u['url']} ({u['status_code']})" for u in err_urls[:15]))

    # Duplicate titles
    dup_titles = sf_data.get("duplicate_titles", [])
    _sf_row(ws_sf, "Nhóm title trùng", len(dup_titles))
    for dt in dup_titles[:20]:
        _sf_row(ws_sf, f"  └ '{dt['title'][:40]}'", len(dt["urls"]),
                " | ".join(dt["urls"][:5]))

    # Missing titles
    mt = sf_data.get("missing_titles", [])
    _sf_row(ws_sf, "Trang thiếu title", len(mt),
            " | ".join(mt[:10]) if mt else "")

    # Missing meta descriptions
    mm = sf_data.get("missing_meta_descriptions", [])
    _sf_row(ws_sf, "Trang thiếu meta description", len(mm),
            " | ".join(mm[:10]) if mm else "")

    # Missing H1
    mh1 = sf_data.get("missing_h1", [])
    _sf_row(ws_sf, "Trang thiếu H1", len(mh1),
            " | ".join(mh1[:10]) if mh1 else "")

    # Images missing alt (None = SF All Crawl Data, no image columns)
    no_alt = sf_data.get("images_missing_alt")
    _sf_row(ws_sf, "Ảnh thiếu alt text",
            len(no_alt) if no_alt is not None else "N/A",
            " | ".join(no_alt[:10]) if no_alt else ("Cần export SF Images tab" if no_alt is None else ""))

    # Non-indexable
    non_ix = sf_data.get("non_indexable", [])
    _sf_row(ws_sf, "Trang non-indexable", len(non_ix))
    reasons_sf: dict[str, int] = {}
    for entry in non_ix:
        r = entry.get("reason", "unknown") or "unknown"
        reasons_sf[r] = reasons_sf.get(r, 0) + 1
    for r, c in sorted(reasons_sf.items(), key=lambda x: -x[1]):
        _sf_row(ws_sf, f"  └ {r}", c)

xl_filename = f"seo_report_{DOMAIN.replace('.', '_')}_{timestamp}.xlsx"
xl_filepath = report_dir / xl_filename
wb.save(xl_filepath)
log_tool("seo_save_excel", 30, xl_filepath.stat().st_size)
print(f"  ✅ Báo cáo Excel: {xl_filepath}")

# ════════════════════════════════════════════════════════════════════════════
# THỐNG KÊ CUỐI
# ════════════════════════════════════════════════════════════════════════════
total_time = time.time() - stats["start_time"]

print(f"\n{'═'*62}")
print("📊 THỐNG KÊ PHIÊN CHẠY")
print(f"{'═'*62}")

print(f"\n  ⏱️  Tổng thời gian: {total_time:.1f}s")
print(f"\n  🕷️  Thu thập dữ liệu:")
fmt("Pages crawled", stats["pages_crawled"])
fmt("Tools called", stats["tools_called"])
fmt("Data thu thập (response)", f"{stats['total_response_bytes']/1024:.0f} KB")
if sf_data:
    fmt("Screaming Frog URLs", f"{sf_data.get('total_urls', 0):,}")
    fmt("SF 4xx/5xx errors", len(sf_data.get("error_urls", [])))
    fmt("SF duplicate title groups", len(sf_data.get("duplicate_titles", [])))
    fmt("SF missing meta desc", len(sf_data.get("missing_meta_descriptions", [])))
    _alt_stat = sf_data.get("images_missing_alt")
    fmt("SF images missing alt", len(_alt_stat) if _alt_stat is not None else "N/A (cần SF Images export)")

print(f"\n  📋 Checklist ({len(TECHNICAL_CHECKLIST)} technical + {len(UI_CHECKLIST)} UI):")
fmt("Tổng tiêu chí", len(all_results))
fmt("✅ Đạt (passed)", audit_score.passed)
fmt("❌ Lỗi (failed)", audit_score.failed)
fmt("⚠️ Cảnh báo (warning)", audit_score.warning)
fmt("🔍 Thủ công (manual)", audit_score.manual)

print(f"\n  🏆 Kết quả đánh giá:")
fmt("Tổng điểm (auto-evaluated)", f"{audit_score.percentage:.1f}%")
fmt("Grade", audit_score.grade)

print(f"\n  📞 Tool calls breakdown:")
for tc in stats["tool_calls"]:
    icon = "✅" if tc["ok"] else "❌"
    print(f"    {icon} {tc['tool']:<42} {tc['ms']:>5}ms  {tc['kb']:>6.1f}KB")

if stats["errors"]:
    print(f"\n  ⚠️  Errors ({len(stats['errors'])}):")
    for e in stats["errors"]:
        print(f"    - {e}")

print(f"\n  🔟 Top 10 Issues:")
priority_label = {"mandatory": "🔴 mandatory", "high": "🟠 high", "nicetohave": "🟡 nicetohave"}
for i, r in enumerate(top_issues, 1):
    icon = "❌" if r.status == "failed" else "⚠️"
    print(f"    {i:2d}. {icon} {priority_label.get(r.priority, r.priority):<16} {r.name}")

print(f"\n{'═'*62}")
print(f"  📄 Report saved: {filepath}")
print(f"{'═'*62}\n")
